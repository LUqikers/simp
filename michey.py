from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook


TOKEN = '1629080653:AAHWNlNRHhpsFfwpyTO26INxAK1WO72K7zg'
book = load_workbook('данные.xlsx')
sheet = book['Лист1']
stickers_page = book['Стикеры']

def main():
    updater = Updater(token=TOKEN)    # объект, который ловит сообщения из телеграм
    dispatcher = updater.dispatcher
    handler = MessageHandler(Filters.all, do_echo)   # отфильтровываем сообщения: теперь устройство должно реагировать
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    sticker_handler = MessageHandler(Filters.sticker, do_stiker)
    keyboard_handler = MessageHandler(Filters.text, do_something)


    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(sticker_handler)
    dispatcher.add_handler(keyboard_handler)
    dispatcher.add_handler(handler)

    updater.start_polling()
    updater.idle()

def do_echo(update: Update, context):
    user = update.message.from_user.is_bot
    name = update.message.from_user.first_name
    if user == True:
        update.message.reply_text(text=f"бот слит!!!")
    else:
        update.message.reply_text(text=f"?? {name} тебе норм?")
        update.message.reply_text(text=":/")
def do_start(update, context):
    keyboard = [
        ["1", "2", "3"],
    ]
    update.message.reply_text(text="Я тебя не знаю", reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))      # текст сообщения, который бот автоматически будет выдавать
def do_help(update, context):
    update.message.reply_text(text='не могу')
    update.message.reply_sticker('CAACAgIAAxkBAANRYEin_F3iky6JmKCGDsT-vkzLozQAAg8BAAIWfGgDOblLeP5BFL0eBA')

def do_something(update, context):
    text = update.message.text

    print(stickers_page.max_row)
    for row in range(2, stickers_page.max_row + 1):
        catch_phrase = stickers_page.cell(row=row, column=4).value
        print(catch_phrase)
        print(text)
        if catch_phrase in text:
            stickers_id = stickers_page.cell(row=row, column=3).value
            update.message.reply_sticker(stickers_id)
            return

        if text == "1":
            update.message.reply_sticker('CAACAgIAAxkBAAN9YFCQOsw7KohKAZwaEMO6rXXsFAADYEcAAuCjggfq4YgmMciicx4E')
        elif text == "2":
            update.message.reply_sticker('CAACAgIAAxkBAAOFYFCQnKGjEvOL9zwlo4YryoFWBGYAAlJHAALgo4IHxKtS9nMp7rQeBA')
        elif text == "3":
            update.message.reply_text(text="сам лох", reply_markup=ReplyKeyboardRemove())
            update.message.reply_sticker('CAACAgIAAxkBAAOIYFCQvyImRBMcP9Ci56c_XBt1vg8AAh4AA3yS8hRACn2o67ys3R4E')

def do_stiker(update: Update, context):
    sticker_id = update.message.sticker.file_id
    update.message.reply_text(sticker_id)
    update.message.reply_sticker(sticker_id)

main()








